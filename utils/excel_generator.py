import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime

class ExcelGenerator:
    def __init__(self, config):
        self.config = config
        
    def generate_sales_report(self, report_data, filename):
        """Generate Excel sales report"""
        os.makedirs(os.path.dirname(filename), exist_ok=True)
        wb = Workbook()
        ws = wb.active
        ws.title = "تقرير المبيعات"
        
        # RTL
        ws.sheet_view.rightToLeft = True
        
        # Header
        ws['A1'] = "تقرير المبيعات"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='right')
        
        ws['A2'] = f"التاريخ: {datetime.now().strftime('%Y/%m/%d %I:%M %p')}"
        
        # Totals
        row = 4
        totals = report_data.get('totals', {})
        
        ws[f'A{row}'] = 'إجمالي المبيعات'
        ws[f'B{row}'] = f"{totals.get('total_sales', 0):.2f} ج.م"
        row += 1
        
        ws[f'A{row}'] = 'عدد الأوردرات'
        ws[f'B{row}'] = totals.get('total_orders', 0)
        row += 1
        
        ws[f'A{row}'] = 'متوسط الأوردر'
        ws[f'B{row}'] = f"{totals.get('avg_order', 0):.2f} ج.م"
        row += 2
        
        # Top items
        ws[f'A{row}'] = 'الأصناف الأكثر مبيعاً'
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        ws[f'A{row}'] = 'الصنف'
        ws[f'B{row}'] = 'الكمية'
        ws[f'C{row}'] = 'الإجمالي'
        
        # Header style
        for col in ['A', 'B', 'C']:
            ws[f'{col}{row}'].font = Font(bold=True)
            ws[f'{col}{row}'].fill = PatternFill(start_color='CCCCCC', fill_type='solid')
        
        row += 1
        
        for item in report_data.get('top_items', []):
            ws[f'A{row}'] = item['name']
            ws[f'B{row}'] = item['qty']
            ws[f'C{row}'] = f"{item['total']:.2f}"
            row += 1
            
        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        
        wb.save(filename)
        return True
